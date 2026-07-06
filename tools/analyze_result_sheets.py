#!/usr/bin/env python3
"""analyze_result_sheets.py — build reviewer-facing summaries from the result sheets.

Reads the seven per-arm CSVs in results/ and writes two self-contained artifacts so a
reviewer can grasp the evaluation without opening the raw sheets:

  results/result_sheets_analysis.xlsx  — one workbook: a cross-benchmark two-axis
      summary, a per-defense summary, a PEINN-vs-undefended preservation table, and a
      verbatim copy of every raw sheet (one tab each).
  results/ANALYSIS.md                  — the same summary tables with a short narrative.

Aggregation: each metric is averaged across the base models a defense is run on
(Vanilla / NeMo / PEINN / Llama Guard: 4 bases; R2D2: zephyr-7B only). Scoring
conventions match the paper (see results/README.md and docs/DATA_CARD.md): a refusal or
deflection is floored on every Ethics instrument; NeMo refuses GGB (shown as n/a).

Usage:  python tools/analyze_result_sheets.py [results_dir]
"""
from __future__ import annotations
import sys, os
import numpy as np
import pandas as pd

DEFENSES = ["Vanilla", "R2D2", "NeMo", "PEINN", "LlamaGuard"]
DEF_LABEL = {"LlamaGuard": "Llama Guard"}


def _mean(df, col):
    s = pd.to_numeric(df[col], errors="coerce")
    return round(s.mean(), 1) if s.notna().any() else np.nan


def load(d):
    R = {}
    for name in ["harmbench_asr", "taxonomy_asr_by_arm", "taxonomy_asr_by_family",
                 "xstest_orr_ucr", "ethics_mfa_wvs_rqi", "morables_accuracy", "ggb_ih_ib"]:
        R[name] = pd.read_csv(os.path.join(d, name + ".csv"))
    return R


def per_defense(R):
    """One row per defense: metric means across the bases it is run on."""
    hb, tx, xs, et, mo, gg = (R["harmbench_asr"], R["taxonomy_asr_by_arm"], R["xstest_orr_ucr"],
                              R["ethics_mfa_wvs_rqi"], R["morables_accuracy"], R["ggb_ih_ib"])
    rows = []
    for dfn in DEFENSES:
        h, t, x, e, m, g = (df[df.defense == dfn] for df in (hb, tx, xs, et, mo, gg))
        n = h.base.nunique()
        rows.append({
            "Defense": DEF_LABEL.get(dfn, dfn),
            "Bases (n)": n,
            "HarmBench ASR % ↓": _mean(h, "asr_mean"),
            "Taxonomy ASR % ↓": _mean(t, "asr_mean"),
            "XSTest ORR % ↓": _mean(x, "orr"),
            "XSTest UCR % ↓": _mean(x, "ucr"),
            "Ethics MFA ↑": _mean(e, "mfa"),
            "Ethics WVS ↑": _mean(e, "wvs"),
            "Ethics RQI ↑": _mean(e, "rqi"),
            "Ethics Composite ↑": _mean(e, "composite"),
            "MORABLES clean % ↑": _mean(m, "clean_acc"),
            "MORABLES gap": _mean(m, "gap"),
            "GGB IH ↓": _mean(g, "ih_mean") if g["ih_mean"].notna().any() else "n/a (refused)",
            "GGB IB ↑": _mean(g, "ib_mean") if g["ib_mean"].notna().any() else "n/a (refused)",
        })
    return pd.DataFrame(rows)


def two_axis(R):
    """Condensed do-no-harm vs moral-reasoning view per defense."""
    pd_ = per_defense(R)
    out = pd_[["Defense", "Bases (n)",
               "HarmBench ASR % ↓", "Taxonomy ASR % ↓",
               "XSTest ORR % ↓", "XSTest UCR % ↓",
               "Ethics Composite ↑", "MORABLES clean % ↑"]].copy()
    return out


def preservation(R):
    """PEINN vs undefended (Vanilla) per base: does routing keep moral reasoning?"""
    et, mo = R["ethics_mfa_wvs_rqi"], R["morables_accuracy"]
    rows = []
    for base in ["zephyr-7B", "qwen2.5-7B", "gemma4-e4B", "gemma3-12B"]:
        ev = et[(et.base == base) & (et.defense == "Vanilla")].iloc[0]
        ep = et[(et.base == base) & (et.defense == "PEINN")].iloc[0]
        mv = mo[(mo.base == base) & (mo.defense == "Vanilla")].iloc[0]
        mp = mo[(mo.base == base) & (mo.defense == "PEINN")].iloc[0]
        rows.append({
            "Base": base,
            "Ethics Comp. Vanilla": ev.composite, "Ethics Comp. PEINN": ep.composite,
            "Δ Comp.": round(ep.composite - ev.composite, 1),
            "MORABLES Vanilla": mv.clean_acc, "MORABLES PEINN": mp.clean_acc,
            "Δ MORABLES": round(mp.clean_acc - mv.clean_acc, 1),
        })
    return pd.DataFrame(rows)


# ---------- Excel ----------
def write_xlsx(path, R, tabs):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E79"); hdr_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=13, color="1F4E79")
    thin = Side(style="thin", color="D9D9D9"); border = Border(thin, thin, thin, thin)

    def sheet(ws, title, df, note=""):
        ws["A1"] = title; ws["A1"].font = title_font
        r0 = 3
        if note:
            ws["A2"] = note; ws["A2"].font = Font(italic=True, size=9, color="595959")
        for j, col in enumerate(df.columns, 1):
            c = ws.cell(r0, j, col); c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = border
        for i, (_, row) in enumerate(df.iterrows(), 1):
            for j, col in enumerate(df.columns, 1):
                v = row[col]
                if isinstance(v, float) and v != v: v = ""
                c = ws.cell(r0 + i, j, v); c.border = border
                c.alignment = Alignment(horizontal="center" if j > 1 else "left")
        for j, col in enumerate(df.columns, 1):
            w = max(len(str(col)) * 0.95, *(len(str(row[col])) for _, row in df.iterrows())) if len(df) else len(str(col))
            ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 9), 22)
        ws.freeze_panes = ws.cell(r0 + 1, 2)

    # ordered tabs: summaries first, then raw sheets
    first = True
    for title, df, note in tabs:
        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]; first = False
        sheet(ws, title, df, note)
    wb.save(path)


# ---------- Markdown ----------
def md_table(df):
    cols = list(df.columns)
    out = ["| " + " | ".join(cols) + " |", "|" + "|".join(["---"] * len(cols)) + "|"]
    for _, row in df.iterrows():
        cells = ["" if (isinstance(row[c], float) and row[c] != row[c]) else str(row[c]) for c in cols]
        out.append("| " + " | ".join(cells) + " |")
    return "\n".join(out)


def write_md(path, two, byd, pres):
    txt = f"""# Result-sheet analysis (PEINN v2.1)

A self-contained summary of the seven per-arm result sheets in this folder, so the
evaluation can be read without opening the raw CSVs. Every number here is derived
directly from those sheets; the same content is in
[`result_sheets_analysis.xlsx`](result_sheets_analysis.xlsx) (one tab per view, plus a
verbatim copy of each raw sheet). Metric direction is marked ↓ (lower is better) or
↑ (higher is better). Values are averaged across the base models each defense is run
on (Vanilla / NeMo / PEINN / Llama Guard: four bases; R2D2: zephyr-7B only).

## 1. Two-axis summary (the headline view)

The evaluation asks two things of a defense at once: does it drive the harm-side
attack success to the floor (**do-no-harm**), and does it keep the base model's moral
reasoning answerable (**moral-reasoning preservation**)? Averaged over the bases:

{md_table(two)}

Reading it: the two routing-based defenses (NeMo, PEINN) both take HarmBench and
jailbreak-taxonomy ASR to the floor, but they diverge on the moral-reasoning axis —
NeMo refuses the moral batteries (Ethics composite and MORABLES fall to the floor),
whereas PEINN keeps them answered and close to the undefended (Vanilla) level, at a
modest rise in XSTest over-refusal. Llama Guard, a second 7B classifier, is competitive
on safety and keeps the batteries scorable, but at that model's deployment cost.

## 2. Per-defense summary (all metrics)

{md_table(byd)}

## 3. Does routing preserve moral reasoning? PEINN vs. undefended, per base

PEINN's second pass should *surface, not manufacture* reasoning: it must not degrade a
competent base model. Comparing PEINN to the undefended Vanilla arm on each base:

{md_table(pres)}

Ethics composite and MORABLES clean accuracy track the undefended baseline on every
base (small, mostly single-digit deltas), consistent with the paper's claim of
non-degradation rather than a leaderboard win.

## Scoring conventions (as in the paper)

- **Ethics** uses a uniform refusal floor: a refusal or deflection (no ratable answer)
  is scored at the instrument floor on the same basis for every arm — an MFQ item at
  maximum error (0 to MFA), a WVS item as out-of-range (0), a refused dilemma at the RQI
  floor. NeMo declines the whole questionnaire and dilemmas, so it floors to ≈ 0.
- **XSTest** counts a NeMo guard-block on a safe prompt as an over-refusal (ORR).
- **MORABLES** uses the deterministic gold-position shuffle; NeMo refuses all items (0).
- **GGB** carries no ground-truth label, so IH/IB are distributional means (no
  correctness claim); NeMo refused every GGB item (shown as n/a).

## Files

`harmbench_asr` · `taxonomy_asr_by_arm` · `taxonomy_asr_by_family` ·
`xstest_orr_ucr` · `ethics_mfa_wvs_rqi` · `morables_accuracy` · `ggb_ih_ib`
(all `.csv` in this folder). Regenerate this analysis with
`python tools/analyze_result_sheets.py`.
"""
    with open(path, "w") as f:
        f.write(txt)


def main():
    d = sys.argv[1] if len(sys.argv) > 1 else "results"
    R = load(d)
    two, byd, pres = two_axis(R), per_defense(R), preservation(R)
    raw_note = "Verbatim copy of the result sheet (per-arm; see paper for full context)."
    tabs = [
        ("Two-axis summary", two, "Do-no-harm vs moral-reasoning preservation, averaged across bases."),
        ("Per-defense summary", byd, "All metrics, averaged across the bases each defense is run on."),
        ("PEINN vs undefended", pres, "Moral-reasoning preservation: PEINN vs Vanilla, per base."),
        ("HarmBench", R["harmbench_asr"], raw_note),
        ("Taxonomy by arm", R["taxonomy_asr_by_arm"], raw_note),
        ("Taxonomy by family", R["taxonomy_asr_by_family"], raw_note),
        ("XSTest", R["xstest_orr_ucr"], raw_note),
        ("Ethics", R["ethics_mfa_wvs_rqi"], raw_note),
        ("MORABLES", R["morables_accuracy"], raw_note),
        ("GGB", R["ggb_ih_ib"], raw_note),
    ]
    write_xlsx(os.path.join(d, "result_sheets_analysis.xlsx"), R, tabs)
    write_md(os.path.join(d, "ANALYSIS.md"), two, byd, pres)
    print("wrote", os.path.join(d, "result_sheets_analysis.xlsx"))
    print("wrote", os.path.join(d, "ANALYSIS.md"))


if __name__ == "__main__":
    main()
